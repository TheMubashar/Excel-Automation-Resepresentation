# ! python 3.7

import os
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def Process_Workbook(FilePath, Filename):
    os.chdir(FilePath)
    wb = xl.load_workbook(Filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 3)
        update_price = cell.value * 0.9
        update_price_cell = sheet.cell(row, 4)
        update_price_cell.value = update_price

    wb.save('Testing2.xlsx')
    values = Reference(sheet,
                       min_row=2,
                       max_row=5,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, row=2, col=5)
    wb.save('Testing2.xlsx')


# User file path and file name
userinput1 = input('Enter your file path: ')
userinput2 = input('Enter your file name: ')

# Adding extension
filename = userinput2 + '.xlsx'
# spliting file path so that we can add \\ to path.
Path = userinput1.split("\\")
filepath = ''
i = 1
lenght = (len(Path))
for x in Path:
    if i == lenght:
        filepath += x
    else:
        i += 1
        filepath += x+'\\'+'\\'

# Workbook call
Process_Workbook(filepath, filename)
